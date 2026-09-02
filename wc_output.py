"""
WildCatcher output / export engine.

Builds the detection report from a list of per-file records, including only the
fields the client selected, in the format(s) they chose. Decoupled from the
processing engine so fields and formats are easy to extend.

Formats:
  csv       - comma-separated, customizable columns
  json      - list of objects, customizable keys
  xlsx      - Excel (File Details + Summary sheets), customizable columns
  sqlite    - generic SQLite .db (one 'detections' table), customizable columns
  timelapse - native Timelapse template (.tdb) + data (.ddb) SQLite files
              (best-effort; see write_timelapse notes)

A "record" is a flat dict keyed by the field keys below; missing keys export
as empty, so the processing layer can populate whatever it has.
"""
import os
import re
import csv
import json
import sqlite3
from datetime import datetime

# Control characters that crash csv.writer ("need to escape, but no escapechar
# set") and openpyxl (IllegalCharacterError). Trail cameras (e.g. BUSHNELL)
# commonly NUL-pad EXIF Make/Model fields, so strip these before writing.
_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _clean_cell(v):
    """Strip control chars from string values; pass everything else through."""
    if isinstance(v, str):
        cleaned = _CONTROL_CHARS.sub("", v)
        return cleaned.rstrip() if cleaned != v else v
    return v

# ---------------------------------------------------------------------------
# Field registry
# ---------------------------------------------------------------------------
# Groups (used to organize the settings GUI)
GROUP_FILE = "file"
GROUP_TIME = "time"
GROUP_IMAGE = "image"
GROUP_VIDEO = "video"
GROUP_DETECTION = "detection"
GROUP_SPECIES = "species"
GROUP_EXIF = "exif"
GROUP_RUN = "run"

GROUP_ORDER = [GROUP_FILE, GROUP_TIME, GROUP_DETECTION, GROUP_SPECIES,
               GROUP_IMAGE, GROUP_VIDEO, GROUP_EXIF, GROUP_RUN]

GROUP_LABELS = {  # English; i18n key is "group_<name>"
    GROUP_FILE: "File", GROUP_TIME: "Time", GROUP_IMAGE: "Image",
    GROUP_VIDEO: "Video", GROUP_DETECTION: "Detection", GROUP_SPECIES: "Species",
    GROUP_EXIF: "Camera / EXIF", GROUP_RUN: "Run info",
}

# (key, stable export header, group, i18n key for the GUI checkbox label)
FIELDS = [
    ("id",                 "ID",                  GROUP_FILE,      "field_id"),
    ("file_name",          "File Name",           GROUP_FILE,      "field_file_name"),
    ("relative_path",      "Relative Path",       GROUP_FILE,      "field_relative_path"),
    ("full_path",          "Full Path",           GROUP_FILE,      "field_full_path"),
    ("folder",             "Folder",              GROUP_FILE,      "field_folder"),
    ("station",            "Station",             GROUP_FILE,      "field_station"),
    ("file_type",          "File Type",           GROUP_FILE,      "field_file_type"),
    ("file_size_kb",       "File Size (KB)",      GROUP_FILE,      "field_file_size"),
    ("file_modified",      "File Modified",       GROUP_FILE,      "field_file_modified"),
    ("time",               "Time",                GROUP_TIME,      "field_time"),
    ("event_id",           "Capture Event",       GROUP_TIME,      "field_event_id"),
    ("detection",          "Detection",           GROUP_DETECTION, "field_detection"),
    ("total_detections",   "Total Detections",    GROUP_DETECTION, "field_total_detections"),
    ("animal_count",       "Animal Count",        GROUP_DETECTION, "field_animal_count"),
    ("human_count",        "Human/Vehicle Count", GROUP_DETECTION, "field_human_count"),
    ("detection_accuracy", "Detection Accuracy",  GROUP_DETECTION, "field_detection_accuracy"),
    ("species",            "Species",             GROUP_SPECIES,   "field_species"),
    ("species_all",        "All Species",         GROUP_SPECIES,   "field_species_all"),
    ("species_counts",     "Species Counts",      GROUP_SPECIES,   "field_species_counts"),
    ("species_accuracy",   "Species Accuracy",    GROUP_SPECIES,   "field_species_accuracy"),
    ("image_width",        "Image Width",         GROUP_IMAGE,     "field_image_width"),
    ("image_height",       "Image Height",        GROUP_IMAGE,     "field_image_height"),
    ("video_length",       "Video Length (s)",    GROUP_VIDEO,     "field_video_length"),
    ("video_fps",          "Video FPS",           GROUP_VIDEO,     "field_video_fps"),
    ("frames_processed",   "Frames Processed",    GROUP_VIDEO,     "field_frames_processed"),
    ("camera_make",        "Camera Make",         GROUP_EXIF,      "field_camera_make"),
    ("camera_model",       "Camera Model",        GROUP_EXIF,      "field_camera_model"),
    ("gps_latitude",       "GPS Latitude",        GROUP_EXIF,      "field_gps_lat"),
    ("gps_longitude",      "GPS Longitude",       GROUP_EXIF,      "field_gps_lon"),
    ("processing_date",    "Processing Date",     GROUP_RUN,       "field_processing_date"),
    ("models_used",        "Models Used",         GROUP_RUN,       "field_models_used"),
    ("processing_notes",   "Processing Notes",    GROUP_RUN,       "field_processing_notes"),
]

FIELD_BY_KEY = {f[0]: f for f in FIELDS}
ALL_FIELD_KEYS = [f[0] for f in FIELDS]

# Back-compat: the original detection_report.xlsx columns + xlsx-only.
DEFAULT_FIELDS = ["id", "file_name", "detection", "species", "time",
                  "video_length", "detection_accuracy", "species_accuracy"]
DEFAULT_FORMATS = ["xlsx"]

ALL_FORMATS = ["csv", "json", "xlsx", "sqlite", "pdf",
               "megadetector", "wildlife_insights", "timelapse"]
FORMAT_LABELS = {  # i18n key is "format_<name>"
    "csv": "CSV (.csv)", "json": "JSON (.json)", "xlsx": "Excel (.xlsx)",
    "sqlite": "SQLite database (.db)",
    "pdf": "PDF report (.pdf)",
    "megadetector": "MegaDetector JSON (.json)",
    "wildlife_insights": "Wildlife Insights CSV (.csv)",
    "timelapse": "Timelapse (.ddb/.tdb)",
}

REPORT_BASENAME = "detection_report"


def header_for(key):
    f = FIELD_BY_KEY.get(key)
    return f[1] if f else key


def fields_by_group():
    """Return {group: [(key, header, i18n_key), ...]} in display order."""
    out = {g: [] for g in GROUP_ORDER}
    for key, header, group, i18n in FIELDS:
        out.setdefault(group, []).append((key, header, i18n))
    return out


def normalize_fields(fields):
    """Drop unknown keys, preserve order; fall back to defaults if empty."""
    valid = [k for k in (fields or []) if k in FIELD_BY_KEY]
    return valid or list(DEFAULT_FIELDS)


def _headers(fields):
    return [header_for(k) for k in fields]


def _row(record, fields):
    return [_clean_cell(record.get(k, "")) for k in fields]


# ---------------------------------------------------------------------------
# Individual format writers
# ---------------------------------------------------------------------------
def write_csv(path, records, fields):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(_headers(fields))
        for r in records:
            w.writerow(_row(r, fields))


def write_json(path, records, fields):
    data = [{k: _clean_cell(r.get(k, "")) for k in fields} for r in records]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


def write_xlsx(path, records, fields, summary=None):
    """
    Two tabs:
      - "Summary"      : whole-batch totals (animals/humans/empty + per species)
      - "File Details" : one row per processed image/video, user-selected columns
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    bold = Font(bold=True)
    wb = Workbook()

    # --- Sheet 1: Summary (batch as a whole) ---
    ws = wb.active
    ws.title = "Summary"

    def _bold_row(*vals):
        ws.append(list(vals))
        for cell in ws[ws.max_row]:
            cell.font = bold

    s = summary or {}
    _bold_row("WildCatcher — Batch Summary")
    ws.append([])
    _bold_row("Metric", "Value")
    ws.append(["Total files found", s.get("total_files", len(records))])
    ws.append(["Files in report", s.get("total_processed", len(records))])
    ws.append(["Animals", s.get("total_animal", 0)])
    ws.append(["Humans / Vehicles", s.get("total_human", 0)])
    ws.append(["Empty (no detection)", s.get("total_empty", 0)])
    if s.get("total_errors"):
        ws.append(["Files skipped (errors)", s.get("total_errors", 0)])
    if s.get("stopped_early"):
        ws.append(["Stopped early by user", "Yes"])

    if s.get("total_events"):
        ws.append(["Capture events", s.get("total_events")])

    ws.append([])
    _bold_row("Species", "Count")
    sc = s.get("species_counts") or {}
    if sc:
        for sp in sorted(sc):
            ws.append([sp, sc[sp]])
    else:
        ws.append(["(no species classified)", 0])

    st = s.get("station_counts") or {}
    if st and not (len(st) == 1 and "" in st):
        ws.append([])
        _bold_row("Station", "Files")
        for name in sorted(st):
            ws.append([name or "(unsorted)", st[name]])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    # --- Sheet 2: File Details (per image/video) ---
    ws2 = wb.create_sheet("File Details")
    ws2.append(_headers(fields))
    for cell in ws2[1]:
        cell.font = bold
    for r in records:
        ws2.append(_row(r, fields))
    if records:
        ws2.freeze_panes = "A2"

    wb.save(path)


def write_sqlite(path, records, fields):
    if os.path.exists(path):
        os.remove(path)
    con = sqlite3.connect(path)
    try:
        cols = ", ".join(f'"{k}" TEXT' for k in fields)
        con.execute(f"CREATE TABLE detections ({cols})")
        placeholders = ",".join("?" * len(fields))
        con.executemany(
            f"INSERT INTO detections VALUES ({placeholders})",
            [tuple(_clean_cell(str(r.get(k, ""))) for k in fields) for r in records],
        )
        con.commit()
    finally:
        con.close()


def write_megadetector(path, records, detector_name=None):
    """
    MegaDetector-compatible results JSON (format 1.3). WildCatcher's detector
    already uses MD categories (1 animal / 2 person / 3 vehicle) and normalized
    [x,y,w,h] bboxes, so this plugs straight into MD-ecosystem tools (Timelapse,
    EcoAssist, AddaxAI, camtrap workflows).
    """
    if detector_name is None:
        detector_name = (records[0].get("models_used") if records else "") or "WildCatcher"

    # Species -> classification category index.
    species_list = []
    for r in records:
        for d in (r.get("detections_raw") or []):
            sp = d.get("species")
            if sp and sp not in species_list:
                species_list.append(sp)
    sp_index = {sp: str(i) for i, sp in enumerate(species_list)}

    images = []
    for r in records:
        rel = (r.get("relative_path") or "").replace("\\", "/")
        fname = r.get("file_name", "")
        file_path = f"{rel}/{fname}" if rel else fname
        dets, max_conf = [], 0.0
        for d in (r.get("detections_raw") or []):
            try:
                conf = float(d.get("conf") or 0)
            except (TypeError, ValueError):
                conf = 0.0
            bbox = d.get("bbox") or [0, 0, 0, 0]
            entry = {"category": str(d.get("category", "1")), "conf": round(conf, 4),
                     "bbox": [round(float(x), 5) for x in bbox]}
            sp, spc = d.get("species"), d.get("species_conf")
            if sp and sp in sp_index:
                try:
                    entry["classifications"] = [[sp_index[sp], round(float(spc), 4)]] if spc is not None else [[sp_index[sp], 1.0]]
                except (TypeError, ValueError):
                    pass
            dets.append(entry)
            max_conf = max(max_conf, conf)
        images.append({"file": file_path, "max_detection_conf": round(max_conf, 4),
                       "detections": dets})

    data = {
        "info": {"format_version": "1.3", "detector": detector_name,
                 "detection_completion_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        "detection_categories": {"1": "animal", "2": "person", "3": "vehicle"},
        "classification_categories": {v: k for k, v in sp_index.items()},
        "images": images,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=1, ensure_ascii=False, default=str)


# Column set aligned to Wildlife Insights / generic camera-trap ingestion.
_WI_COLUMNS = ["deployment_id", "location", "filename", "timestamp",
               "common_name", "count", "cv_confidence", "is_blank",
               "latitude", "longitude"]


def write_wildlife_insights(path, records):
    """
    Camera-trap CSV in a Wildlife-Insights-style layout (one row per file).
    A practical starting point for WI / Camtrap-DP ingestion; column mapping may
    need tweaking to your WI project template.
    """
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(_WI_COLUMNS)
        for r in records:
            station = r.get("station") or r.get("relative_path") or r.get("folder") or ""
            is_blank = "true" if (r.get("detection", "") in ("", "empty")) else "false"
            conf = r.get("species_accuracy") or r.get("detection_accuracy") or ""
            w.writerow([_clean_cell(v) for v in (
                station, station, r.get("file_name", ""), r.get("time", ""),
                r.get("species", ""), r.get("animal_count", ""), conf, is_blank,
                r.get("gps_latitude", ""), r.get("gps_longitude", ""),
            )])


def _pdf_safe(text):
    """fpdf2 core fonts are latin-1 only; replace unrepresentable chars so a
    Japanese station/licensee name never crashes the PDF (shows as '?')."""
    return str(text).encode("latin-1", "replace").decode("latin-1")


def write_pdf(path, records, summary=None):
    """A polished one-file PDF summary report (charts + per-station + GPS map).
    Uses fpdf2 (pure-Python, no heavy chart deps)."""
    from fpdf import FPDF

    s = summary or {}
    GREEN = (0x43, 0x78, 0x20)
    LIME = (0x9b, 0xc4, 0x72)
    GREY = (90, 90, 90)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()
    usable_w = pdf.w - pdf.l_margin - pdf.r_margin

    def line(h=6):
        pdf.ln(h)

    def heading(text):
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*GREEN)
        pdf.cell(0, 8, _pdf_safe(text))
        pdf.ln(9)
        pdf.set_text_color(30, 30, 30)

    # --- Title ---
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 12, "WildCatcher - Wildlife Detection Report")
    pdf.ln(12)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*GREY)
    pdf.cell(0, 6, _pdf_safe("Generated " + datetime.now().strftime("%Y-%m-%d %H:%M")))
    pdf.ln(6)
    if s.get("stopped_early"):
        pdf.set_text_color(180, 80, 0)
        pdf.cell(0, 6, "Note: run was stopped early; totals are partial.")
        pdf.ln(6)

    # --- Summary metrics ---
    heading("Summary")
    metrics = [
        ("Files processed", s.get("total_processed", len(records))),
        ("Animals", s.get("total_animal", 0)),
        ("Humans / Vehicles", s.get("total_human", 0)),
        ("Empty (no detection)", s.get("total_empty", 0)),
        ("Capture events", s.get("total_events", 0)),
        ("Stations", len({r.get("station", "") for r in records})),
    ]
    if s.get("total_errors"):
        metrics.append(("Files skipped (errors)", s.get("total_errors")))
    pdf.set_font("Helvetica", "", 11)
    for label, val in metrics:
        pdf.set_text_color(60, 60, 60)
        pdf.cell(70, 7, _pdf_safe(label))
        pdf.set_text_color(20, 20, 20)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, _pdf_safe(val))
        pdf.set_font("Helvetica", "", 11)
        pdf.ln(7)

    # --- Species distribution bar chart ---
    sc = s.get("species_counts") or {}
    if sc:
        heading("Species distribution")
        pdf.set_font("Helvetica", "", 10)
        maxc = max(sc.values()) or 1
        label_w, num_w = 46.0, 16.0
        bar_max = usable_w - label_w - num_w
        x0 = pdf.l_margin + label_w
        for sp in sorted(sc, key=lambda k: -sc[k]):
            y = pdf.get_y()
            pdf.set_text_color(40, 40, 40)
            pdf.cell(label_w, 7, _pdf_safe(sp)[:24])
            bw = (sc[sp] / maxc) * bar_max
            pdf.set_fill_color(*LIME)
            pdf.rect(x0, y + 1.4, max(0.4, bw), 4.4, style="F")
            pdf.set_xy(x0 + bw + 2, y)
            pdf.cell(num_w, 7, str(sc[sp]))
            pdf.ln(7)

    # --- Per-station table ---
    st = s.get("station_counts") or {}
    st = {k: v for k, v in st.items() if k and k != "(unsorted)"}
    if st:
        heading("Files per station")
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(*GREEN)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(usable_w * 0.6, 7, "  Station", border=0, fill=True)
        pdf.cell(usable_w * 0.4, 7, "  Files", border=0, fill=True)
        pdf.ln(7)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(30, 30, 30)
        for i, name in enumerate(sorted(st)):
            if i % 2 == 0:
                pdf.set_fill_color(238, 242, 235)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.cell(usable_w * 0.6, 6.5, "  " + _pdf_safe(name)[:40], fill=True)
            pdf.cell(usable_w * 0.4, 6.5, "  " + str(st[name]), fill=True)
            pdf.ln(6.5)

    # --- GPS map (simple scatter) ---
    pts = []
    for r in records:
        try:
            lon = float(r.get("gps_longitude"))
            lat = float(r.get("gps_latitude"))
        except (TypeError, ValueError):
            continue
        if lon == 0 and lat == 0:
            continue
        pts.append((lon, lat))
    if pts:
        heading("Detection locations (GPS)")
        map_h = 70.0
        x0, y0 = pdf.l_margin, pdf.get_y()
        pdf.set_draw_color(*GREEN)
        pdf.set_fill_color(245, 248, 243)
        pdf.rect(x0, y0, usable_w, map_h, style="DF")
        lons = [p[0] for p in pts]
        lats = [p[1] for p in pts]
        lon_min, lon_max = min(lons), max(lons)
        lat_min, lat_max = min(lats), max(lats)
        lon_span = (lon_max - lon_min) or 1e-6
        lat_span = (lat_max - lat_min) or 1e-6
        pad = 5.0
        pdf.set_fill_color(*GREEN)
        for lon, lat in pts:
            px = x0 + pad + (lon - lon_min) / lon_span * (usable_w - 2 * pad)
            # invert lat so north is up
            py = y0 + pad + (lat_max - lat) / lat_span * (map_h - 2 * pad)
            pdf.ellipse(px - 0.9, py - 0.9, 1.8, 1.8, style="F")
        pdf.set_y(y0 + map_h + 1)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*GREY)
        pdf.cell(0, 5, _pdf_safe(
            f"Lat {lat_min:.4f} to {lat_max:.4f}, Lon {lon_min:.4f} to {lon_max:.4f}  "
            f"({len(pts)} located)"))
        pdf.ln(6)

    # --- Footer ---
    try:
        import wc_version
        ver = wc_version.APP_VERSION
    except Exception:
        ver = ""
    pdf.set_y(-14)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(*GREY)
    pdf.cell(0, 6, _pdf_safe(f"Generated by WildCatcher {ver}"), align="C")

    pdf.output(path)


def write_timelapse(out_dir, records):
    """
    Best-effort native Timelapse export: a template (.tdb) and a data (.ddb)
    SQLite database, one DataTable row per file.

    NOTE: Timelapse's schema is version-specific and cannot be validated without
    Timelapse installed. This targets the common Timelapse 2.x layout. To
    guarantee compatibility, replace TEMPLATE_CONTROLS / table creation with the
    schema read from a sample .tdb produced by your Timelapse version, or open
    the output in Timelapse and adjust. Kept isolated here for easy fixes.
    """
    tdb_path = os.path.join(out_dir, REPORT_BASENAME + ".tdb")
    ddb_path = os.path.join(out_dir, REPORT_BASENAME + ".ddb")

    # Standard controls + WildCatcher data fields (DataLabel -> column).
    # (Type, DataLabel, Label, DefaultValue, Tooltip, List)
    controls = [
        ("File",         "File",         "File",         "", "File name", ""),
        ("RelativePath", "RelativePath", "RelativePath", "", "Relative path", ""),
        ("DateTime",     "DateTime",     "DateTime",     "", "Capture time", ""),
        ("DeleteFlag",   "DeleteFlag",   "Delete?",      "false", "Mark for deletion", ""),
        ("Note",         "Species",      "Species",      "", "Detected species", ""),
        ("Counter",      "Count",        "Count",        "0", "Animal count", ""),
        ("Note",         "Confidence",   "Confidence",   "", "Detection confidence", ""),
        ("FixedChoice",  "Detection",    "Detection",    "", "animal/human/empty",
         "animal|human|empty"),
    ]
    data_labels = [c[1] for c in controls]

    def _build_template(con):
        con.execute(
            "CREATE TABLE TemplateTable ("
            "Id INTEGER PRIMARY KEY AUTOINCREMENT, ControlOrder INTEGER, "
            "SpreadsheetOrder INTEGER, Type TEXT, DefaultValue TEXT, Label TEXT, "
            "DataLabel TEXT, Tooltip TEXT, Width INTEGER, Copyable TEXT, "
            "Visible TEXT, List TEXT)"
        )
        for i, (ctype, dlabel, label, default, tip, lst) in enumerate(controls, start=1):
            con.execute(
                "INSERT INTO TemplateTable (ControlOrder, SpreadsheetOrder, Type, "
                "DefaultValue, Label, DataLabel, Tooltip, Width, Copyable, Visible, List) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (i, i, ctype, default, label, dlabel, tip, 100, "true", "true", lst),
            )

    # --- .tdb (template only) ---
    if os.path.exists(tdb_path):
        os.remove(tdb_path)
    con = sqlite3.connect(tdb_path)
    try:
        _build_template(con)
        con.commit()
    finally:
        con.close()

    # --- .ddb (template copy + ImageSetTable + DataTable) ---
    if os.path.exists(ddb_path):
        os.remove(ddb_path)
    con = sqlite3.connect(ddb_path)
    try:
        _build_template(con)
        con.execute(
            "CREATE TABLE ImageSetTable ("
            "Id INTEGER PRIMARY KEY AUTOINCREMENT, Log TEXT, Row INTEGER, "
            "Selection INTEGER, TimeZone TEXT, VersionCompatibility TEXT, "
            "SortTerms TEXT, WhiteSpaceTrimmed TEXT)"
        )
        con.execute(
            "INSERT INTO ImageSetTable (Log, Row, Selection, TimeZone, "
            "VersionCompatibility, SortTerms, WhiteSpaceTrimmed) VALUES "
            "('Generated by WildCatcher', 0, 0, '', '2.3.0.0', '', 'true')"
        )
        cols = ", ".join(f'"{c}" TEXT' for c in data_labels)
        con.execute(f"CREATE TABLE DataTable (Id INTEGER PRIMARY KEY AUTOINCREMENT, {cols})")
        placeholders = ",".join("?" * len(data_labels))
        rows = []
        for r in records:
            rows.append(tuple(_clean_cell(v) for v in (
                r.get("file_name", ""),
                r.get("relative_path", "") or r.get("folder", ""),
                r.get("time", ""),
                "false",
                r.get("species", ""),
                str(r.get("animal_count", "") or 0),
                str(r.get("detection_accuracy", "")),
                r.get("detection", "") or "empty",
            )))
        con.executemany(
            f'INSERT INTO DataTable ({", ".join(data_labels)}) VALUES ({placeholders})',
            rows,
        )
        con.commit()
    finally:
        con.close()

    return [tdb_path, ddb_path]


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------
def write_reports(records, out_dir, fields=None, formats=None, summary=None, log=None):
    """
    Write the report in each requested format. Returns list of written paths.
    `fields` = ordered list of field keys; `formats` = subset of ALL_FORMATS.
    Falls back to the back-compat defaults (xlsx, original columns). A failure in
    one format is logged and skipped so the others still get written.
    """
    fields = normalize_fields(fields)
    formats = [f for f in (formats or []) if f in ALL_FORMATS] or list(DEFAULT_FORMATS)
    log = log or (lambda m: None)

    # Auto-fill run-level fields if selected
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, r in enumerate(records, start=1):
        if "id" in fields and not r.get("id"):
            r["id"] = i
        if "processing_date" in fields and not r.get("processing_date"):
            r["processing_date"] = now

    written = []
    base = os.path.join(out_dir, REPORT_BASENAME)
    for fmt in formats:
        try:
            if fmt == "csv":
                written.append(_write_locked_safe(
                    base + ".csv", lambda p: write_csv(p, records, fields), log))
            elif fmt == "json":
                written.append(_write_locked_safe(
                    base + ".json", lambda p: write_json(p, records, fields), log))
            elif fmt == "xlsx":
                written.append(_write_locked_safe(
                    base + ".xlsx", lambda p: write_xlsx(p, records, fields, summary), log))
            elif fmt == "sqlite":
                written.append(_write_locked_safe(
                    base + ".db", lambda p: write_sqlite(p, records, fields), log))
            elif fmt == "pdf":
                written.append(_write_locked_safe(
                    base + ".pdf", lambda p: write_pdf(p, records, summary), log))
            elif fmt == "megadetector":
                written.append(_write_locked_safe(
                    base + "_megadetector.json",
                    lambda p: write_megadetector(p, records), log))
            elif fmt == "wildlife_insights":
                written.append(_write_locked_safe(
                    base + "_wildlife_insights.csv",
                    lambda p: write_wildlife_insights(p, records), log))
            elif fmt == "timelapse":
                written.extend(write_timelapse(out_dir, records))
        except Exception as e:
            log(f"  {fmt} export failed: {e}")
    return written


def _write_locked_safe(path, writer, log):
    """
    Write via writer(path). If the destination is locked (e.g. the previous
    detection_report.xlsx is still open in Excel) retry once with a timestamped
    name so a run's results are never silently lost.
    """
    try:
        writer(path)
        return path
    except (PermissionError, OSError) as e:
        base, ext = os.path.splitext(path)
        alt = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        log(f"  {os.path.basename(path)} is in use ({e}); "
            f"saving as {os.path.basename(alt)} instead. "
            "Close it in Excel to overwrite the main file next time.")
        writer(alt)
        return alt
